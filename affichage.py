# -*- coding: utf-8 -*-
"""
Created on Fri May 14 11:59:06 2021

@author: mar_altermark
"""
import xlrd
import sys
import openpyxl
import random
import fonctions as f
from openpyxl.styles import Alignment, Font
from openpyxl.chart import BarChart,PieChart,Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import PatternFill
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as Ft


from openpyxl.styles.colors import Color
def enregistre_resultat(produits_marque):
    
    nom = "BC produits - test"+str(int(100000*random.random()))+".xlsx"
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    fichierResultat = openpyxl.Workbook()
    feuille_principale = fichierResultat.active
    #ONGLET PRINCIPAL =============
    nbproduits = 0
    feuille_principale.title= "Résultat général" 
    feuille_principale.append(produits_marque[0][1][0][0:4]+produits_marque[0][1][0][6:])
    for marque in produits_marque[1:]:
        for produit in marque[1]:
            resultat = [produit[i] for i in [0,1,2,3]]+produit[6]
            nbproduits+=1
            feuille_principale.append(resultat)
    for row in feuille_principale.iter_rows(min_row=2):
        for cell in row[:13]:
            cell.number_format = '0.0'
        for cell in row[13:15]:
            cell.number_format = '0.0%'
    tailles = [25,10,12,8,12,12,12,10,13,13,13,12,13,10,10,5,5,3]
    feuille_principale.row_dimensions[1].height = 45
    for dim in range(len(tailles)):        
        feuille_principale[alphabet[dim]+'1'].alignment = Alignment(wrapText=True)
        feuille_principale[alphabet[dim]+'1'].font = Font(bold=True)
        feuille_principale.column_dimensions[alphabet[dim]].width = tailles[dim]
    
    enphaseColonne(feuille_principale, 11)
    enphaseColonne(feuille_principale, 12)
    
    graphique = BarChart()
    graphique.style = 2
    graphique.overlap = 100
    graphique.title = "Emissions des produits par sac"
    graphique.y_axis.title = 'Bilan Carbone (kgCO2e)'
    graphique.type = "bar"
    graphique.x_axis.scaling.orientation = "maxMin"
    graphique.y_axis.minorTickMark = 'in'
    graphique.height = 5.3+ (nbproduits)*10.9/15
    graphique.grouping = "stacked"
    references = Reference(feuille_principale, min_col = 1, min_row = 2, max_row = nbproduits+1)
    for i in [4,5,6,7,8,9,10]:
        #On ajoute une à une les séries
        valeurs =  Reference(feuille_principale, min_col=i+1, min_row=1, max_row=nbproduits+1)
        graphique.add_data(valeurs, titles_from_data=True)
    graphique.set_categories(references)
    feuille_principale.add_chart(graphique, "S1")
    graphique2 =  BarChart()
    graphique2.style = 2
    graphique2.x_axis.scaling.orientation = "maxMin"
    graphique2.overlap = 100
    graphique2.title = "Emissions des produits au m3"
    graphique2.y_axis.title = 'Bilan Carbone (kgCO2e/m3)'
    graphique2.y_axis.minorTickMark = 'in'
    graphique2.type = "bar"
    references = Reference(feuille_principale, min_col = 1, min_row = 2, max_row = nbproduits+1)
    valeurs =  Reference(feuille_principale, min_col=13, min_row=1, max_row=nbproduits+1)
    graphique2.add_data(valeurs, titles_from_data=True)
    graphique2.set_categories(references)
    graphique2.x_axis.min = 0
    graphique2.x_axis.max = 300
    graphique2.y_axis.min = 0
    graphique2.y_axis.max = 300
    graphique2.height = 5.3+ (nbproduits)*9.5/15
    feuille_principale.add_chart(graphique2, "AB1")
    feuille_principale.freeze_panes = 'B2'
    
    
    #ONGLETS PAR MARQUE
    cp = CharacterProperties(sz=800)
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    for marque in produits_marque[1:]:
        currentline = 1
        feuille_secondaire = fichierResultat.create_sheet("BC "+marque[0],1)
        feuille_secondaire.append(["PF","PSF","Marque","Volume",
                                   "Composition","Quantité","Unité",
                                   "Total routier (kgCO2e/sac)","Total naval (kgCO2e/sac)",
                                   "Total fabrication (kgCO2e/sac)","Total N2O (kgCO2e/sac)",
                                   "Total tourbe décomposition (kgCO2e/sac)",
                                   "Total sacherie (kgCO2e/sac)","Emissions fixes (kgCO2e/sac)",
                                   "Bilan carbone par sac (kgCO2e/sac)",
                                   "Bilan carbone par m3 (kgCO2e/m3)",
                                   "Part de la tourbe dans le BC",
                                   "Part des engrais dans le BC"])
        currentline +=1
        for produit in marque[1]:
            taillepdt = 0
            feuille_secondaire.append([produit[i] for i in [0,1,2,3]])
            currentline +=1
            for c in produit[4][1:]:
                feuille_secondaire.append(4*[""]+c[0:2]+["%"]+c[2:])
                currentline +=1
                taillepdt +=1
            for c in produit[5][1:]:
                feuille_secondaire.append(4*[""]+c[0:2]+["kg/m3"]+c[2:])
                currentline +=1
                taillepdt +=1
            while taillepdt<8:
                feuille_secondaire.append(1*[""])
                taillepdt+=1
                currentline +=1
            feuille_secondaire.append(7*[""]+produit[6])
            currentline +=1
            taillepdt +=1
            
            pie = PieChart()
            labels = Reference(feuille_secondaire, min_col=8, min_row=1, max_col=14)
            data = Reference(feuille_secondaire, min_col=8, min_row=currentline-1, max_col = 14)
            pie.add_data(data, from_rows = True, titles_from_data=False)
            pie.set_categories(labels)
            pie.title = feuille_secondaire["A"+str(currentline-taillepdt-1)].value
            pie.legend.position = "tr"
            pie.layout = Layout(manualLayout=ManualLayout(
                x = 0, y = 0,
                h = 1, w = 0.4
                )
                )
            pie.legend.layout = Layout(manualLayout=ManualLayout(
                x = 0.6, y = 0,
                h=1, w=0.4
                )
                )
            
            pie.width = 10.45
            pie.height = (taillepdt-1)*0.56
            
            pie.legend.textProperties = rtp
            feuille_secondaire.add_chart(pie, "A"+str(currentline-taillepdt))
            
        feuille_secondaire.freeze_panes = 'B2'
        tailles = [25,10,12,8,30,7,9,10,11,10,11,13,13,14,13,12,13,12]
        for dim in range(len(tailles)):
            feuille_secondaire[alphabet[dim]+'1'].alignment = Alignment(wrapText=True)
            feuille_secondaire[alphabet[dim]+'1'].font = Font(bold=True)
            feuille_secondaire.column_dimensions[alphabet[dim]].width = tailles[dim]
        feuille_secondaire.row_dimensions[1].height = 45
        enphaseColonne(feuille_secondaire, 15)
        enphaseColonne(feuille_secondaire, 16)
    try:
        fichierResultat.save("Res/"+nom)
    except PermissionError:     #Soit à cause du fichier réseau, soit parce que l'excel est ouvert ailleurs
        print("Impossible d'enregistrer")
    except FileNotFoundError: #Emplacement inexistant
        print("Impossible d'enregistrer")
    return nom

def enphaseColonne(feuille, colonne):
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    n = feuille.max_row
    for i in range(1,n+1):
        feuille[alphabet[colonne]+str(i)].font = Font(bold=True)
    return None

def noterProduits(nom):
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    couleurs = ["0099CC00", "00FFCC00", "00FF9900", "00FF6600", "00800000"] #Du vert au rouge
    try:
         classeur = openpyxl.load_workbook("Res/"+nom)
    except FileNotFoundError:
         sys.exit("Le document résultat n'est pas trouvé à l'emplacement "+ "Res/"+nom)
    feuillePrincipale = classeur.active
    maximumBCsac = 0
    maximumBCm3 = 0
    for row in feuillePrincipale.iter_rows(min_row=2, min_col = 12,max_col=13, max_row=feuillePrincipale.max_row): 
        maximumBCsac = max(row[0].value, maximumBCsac)
        maximumBCm3  = max(row[1].value, maximumBCm3 )
    feuillePrincipale.cell(row=1, column=16, value="Note (par sac)")
    feuillePrincipale.cell(row=1, column=17, value="Note (par m3)")
    for row in feuillePrincipale.iter_rows(min_row=2, min_col = 12,max_col=17, max_row=feuillePrincipale.max_row): 
        parsac = row[0].value
        note1 = int(4-4*(maximumBCsac-parsac)/maximumBCsac)
        parm3 = row[1].value
        note2 = int(4-4*(maximumBCm3- parm3)/ maximumBCm3)
        note_sac = alphabet[note1]
        note_m3  = alphabet[note2]
        row[4].value = note_sac
        row[4].fill = PatternFill(fill_type="solid",start_color=couleurs[note1]) 
        row[5].value = note_m3
        row[5].fill = PatternFill(fill_type="solid",start_color=couleurs[note2]) 
    
    try:
        classeur.save("Res/"+nom)
    except PermissionError:     #Soit à cause du fichier réseau, soit parce que l'excel est ouvert ailleurs
        print("Impossible d'enregistrer")
    except FileNotFoundError: #Emplacement inexistant
        print("Impossible d'enregistrer")
    return nom

def listeUnique(liste):
    result =[]
    for x in liste:
        if x not in result:
            result.append(x)
    return result

def elements_manquants(nom):
    SFM = listeUnique(f.SFMANQUANT)
    SACHM = listeUnique(f.SACHERIEMANQUANTE)
    INTM = listeUnique(f.INTRANTMANQUANT)
    try:
         classeur = openpyxl.load_workbook("Res/"+nom)
    except FileNotFoundError:
         sys.exit("Le document résultat n'est pas trouvé à l'emplacement "+ "Res/"+nom)
    feuille_secondaire = classeur.create_sheet("ELEMENTS MANQUANTS",1)
    feuille_secondaire.append(["Semi-finis", "Sacherie", "Intrants"])
    for i in range(len(SFM)):
        feuille_secondaire["A"+str(i+2)] = SFM[i]
    for i in range(len(SACHM)):
        feuille_secondaire["B"+str(i+2)] = SACHM[i]
    for i in range(len(INTM)):
        feuille_secondaire["C"+str(i+2)] = INTM[i]
    feuille_secondaire.column_dimensions["A"].width = 20
    feuille_secondaire.column_dimensions["B"].width = 20
    feuille_secondaire.column_dimensions["C"].width = 20
    try:
        classeur.save("Res/"+nom)
    except PermissionError:     #Soit à cause du fichier réseau, soit parce que l'excel est ouvert ailleurs
        print("Impossible d'enregistrer")
    except FileNotFoundError: #Emplacement inexistant
        print("Impossible d'enregistrer")
    return nom